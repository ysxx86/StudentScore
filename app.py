from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max-limit

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def get_excellent_threshold(grade):
    if grade in [1, 2]:
        return 90
    elif grade in [3, 4]:
        return 85
    elif grade in [5, 6]:
        return 80
    return 90  # 默认值

def analyze_scores(file_path, grade):
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 查找成绩列
        score_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in ['成绩', '分数', 'score', 'Score']:
                score_col = col
                break
        
        if score_col is None:
            return {'error': '未找到成绩列'}
        
        # 收集所有成绩
        scores = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=score_col).value
            try:
                score = float(cell_value) if cell_value is not None else None
                if score is not None:
                    scores.append(score)
            except (ValueError, TypeError):
                continue
        
        if not scores:
            return {'error': '没有找到有效的成绩数据'}
        
        # 统计各分数段人数
        ranges = {
            '90-100': len([s for s in scores if 90 <= s <= 100]),
            '80-89': len([s for s in scores if 80 <= s < 90]),
            '70-79': len([s for s in scores if 70 <= s < 80]),
            '60-69': len([s for s in scores if 60 <= s < 70]),
            '60以下': len([s for s in scores if s < 60])
        }
        
        # 计算其他统计数据
        total_students = len(scores)
        total_score = sum(scores)
        average_score = total_score / total_students if total_students > 0 else 0
        excellent_threshold = get_excellent_threshold(grade)
        excellent_count = len([s for s in scores if s >= excellent_threshold])
        pass_count = len([s for s in scores if s >= 60])
        
        return {
            'ranges': ranges,
            'total_score': round(float(total_score), 2),
            'average_score': round(float(average_score), 2),
            'excellent_rate': round(excellent_count / total_students * 100, 2),
            'pass_rate': round(pass_count / total_students * 100, 2),
            'excellent_count': excellent_count,
            'total_students': total_students
        }
    except Exception as e:
        return {'error': f'分析成绩时发生错误：{str(e)}'}

@app.route('/')
def index():
    return send_file('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件被上传'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'})
        
        # 检查文件大小
        if request.content_length > app.config['MAX_CONTENT_LENGTH']:
            return jsonify({'error': f'文件大小超过限制（最大{app.config["MAX_CONTENT_LENGTH"]//1024//1024}MB）'})
        
        # 检查文件类型
        if not file.filename.lower().endswith(('.xls', '.xlsx')):
            return jsonify({'error': '请上传Excel文件（.xls或.xlsx格式）'})
        
        grade = request.form.get('grade', type=int)
        if not grade or grade not in range(1, 7):
            return jsonify({'error': '请选择正确的年级（1-6年级）'})
        
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            file.save(file_path)
        except IOError as e:
            return jsonify({'error': f'保存文件时发生错误：{str(e)}'})
        
        try:
            result = analyze_scores(file_path, grade)
            if 'error' in result:
                return jsonify({'error': result['error']})
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': f'分析成绩时发生错误：{str(e)}'})
        finally:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass  # 忽略删除临时文件时的错误
                    
    except Exception as e:
        return jsonify({'error': f'处理上传请求时发生错误：{str(e)}'})

if __name__ == '__main__':
    app.run(debug=True)